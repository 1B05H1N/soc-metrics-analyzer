#!/usr/bin/env python3
"""
Enhanced Excel Report Generator
Creates professional Excel reports with charts, tables, and visualizations
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from typing import Dict, List, Any
from config import Config

class ExcelReportGenerator:
    """Enhanced Excel report generator with professional formatting"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
        
        # Color scheme for professional appearance
        self.colors = {
            'header': '366092',      # Dark blue
            'subheader': '4472C4',   # Medium blue
            'accent': '70AD47',      # Green
            'warning': 'FFC000',     # Orange
            'error': 'C00000',       # Red
            'light_gray': 'F2F2F2',  # Light gray
            'border': 'D0D0D0'       # Border gray
        }
        
        # Styles
        self.styles = {
            'header_font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
            'subheader_font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'normal_font': Font(name='Calibri', size=11),
            'bold_font': Font(name='Calibri', size=11, bold=True),
            'small_font': Font(name='Calibri', size=10),
            'header_fill': PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid'),
            'subheader_fill': PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid'),
            'accent_fill': PatternFill(start_color=self.colors['accent'], end_color=self.colors['accent'], fill_type='solid'),
            'light_fill': PatternFill(start_color=self.colors['light_gray'], end_color=self.colors['light_gray'], fill_type='solid'),
            'border': Border(
                left=Side(style='thin', color=self.colors['border']),
                right=Side(style='thin', color=self.colors['border']),
                top=Side(style='thin', color=self.colors['border']),
                bottom=Side(style='thin', color=self.colors['border'])
            ),
            'center_alignment': Alignment(horizontal='center', vertical='center'),
            'left_alignment': Alignment(horizontal='left', vertical='center'),
            'right_alignment': Alignment(horizontal='right', vertical='center')
        }
    
    def create_report(self, data: Dict[str, Any], filename: str) -> bool:
        """Create comprehensive Excel report"""
        try:
            self.wb = Workbook()
            
            # Create all sheets
            self._create_summary_sheet(data)
            self._create_metrics_sheet(data)
            self._create_trends_sheet(data)
            self._create_breakdown_sheet(data)
            self._create_performance_sheet(data)
            self._create_sla_sheet(data)
            self._create_raw_data_sheet(data)
            
            # Save workbook
            self.wb.save(filename)
            return True
            
        except Exception as e:
            print(f"ERROR: Failed to create Excel report: {e}")
            return False
    
    def _create_summary_sheet(self, data: Dict[str, Any]):
        """Create executive summary sheet"""
        ws = self.wb.active
        ws.title = Config.EXCEL_CONFIG['SHEETS']['SUMMARY']
        
        # Title
        ws['A1'] = 'SOC Metrics Executive Summary'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Report metadata
        row = 3
        ws[f'A{row}'] = 'Report Period:'
        ws[f'B{row}'] = data.get('analysis_period', 'Last 30 days')
        ws[f'A{row+1}'] = 'Generated:'
        ws[f'B{row+1}'] = data.get('generated_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ws[f'A{row+2}'] = 'Analysis Type:'
        ws[f'B{row+2}'] = data.get('analysis_name', 'All Tickets Analysis')
        
        # Key metrics table
        row = 8
        self._create_metrics_table(ws, data, row)
        
        # Performance scores
        row = 15
        self._create_performance_scores(ws, data, row)
        
        # Recommendations
        row = 22
        self._create_recommendations(ws, data, row)
    
    def _safe_excel_value(self, value):
        """Convert value to Excel-compatible format"""
        if isinstance(value, list):
            return ', '.join(str(item) for item in value)
        elif isinstance(value, dict):
            return str(value)
        elif value is None:
            return ''
        else:
            return str(value)
    
    def _create_metrics_table(self, ws, data: Dict[str, Any], start_row: int):
        """Create key metrics table"""
        try:
            # Headers
            headers = ['Metric', 'Value', 'Target', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = self.styles['subheader_font']
                cell.fill = self.styles['subheader_fill']
                cell.border = self.styles['border']
                cell.alignment = self.styles['center_alignment']
            
            # Safely extract values with defaults
            mttr_data = data.get('mttr', {})
            mtd_data = data.get('mtd', {})
            
            mttr_hours = mttr_data.get('mttr_hours', 0) if isinstance(mttr_data, dict) else 0
            mtd_hours = mtd_data.get('mtd_hours', 0) if isinstance(mtd_data, dict) else 0
            sla_breaches = data.get('sla_breaches', 0)
            total_tickets = data.get('total_tickets', 0)
            
            # Data
            metrics_data = [
                ['Total Tickets', total_tickets, 'N/A', ''],
                ['MTTR (Hours)', f"{mttr_hours:.2f}", '4.0', self._get_status_color(mttr_hours, 4.0)],
                ['MTD (Hours)', f"{mtd_hours:.2f}", '1.0', self._get_status_color(mtd_hours, 1.0)],
                ['SLA Breaches', sla_breaches, '0', self._get_status_color(sla_breaches, 0, reverse=True)]
            ]
            
            for i, (metric, value, target, status) in enumerate(metrics_data, 1):
                row = start_row + i
                ws.cell(row=row, column=1, value=metric).font = self.styles['bold_font']
                ws.cell(row=row, column=2, value=value).font = self.styles['normal_font']
                ws.cell(row=row, column=3, value=target).font = self.styles['normal_font']
                ws.cell(row=row, column=4, value=status).font = self.styles['normal_font']
                
                # Apply borders
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.styles['border']
                    
        except Exception as e:
            print(f"ERROR: Failed to create metrics table: {e}")
            # Create a simple fallback table
            ws.cell(row=start_row, column=1, value='Metric').font = self.styles['bold_font']
            ws.cell(row=start_row, column=2, value='Value').font = self.styles['bold_font']
            ws.cell(row=start_row+1, column=1, value='Total Tickets').font = self.styles['normal_font']
            ws.cell(row=start_row+1, column=2, value=str(data.get('total_tickets', 'N/A'))).font = self.styles['normal_font']
    
    def _create_performance_scores(self, ws, data: Dict[str, Any], start_row: int):
        """Create performance scores section"""
        try:
            ws.cell(row=start_row, column=1, value='Performance Scores').font = Font(name='Calibri', size=14, bold=True)
            
            # Safely extract MTTR data
            mttr_data = data.get('mttr', {})
            mttr_hours = mttr_data.get('mttr_hours', 0) if isinstance(mttr_data, dict) else 0
            mttr_score = self._calculate_performance_score(mttr_hours, Config.PERFORMANCE_THRESHOLDS['MTTR'])
            
            ws.cell(row=start_row+2, column=1, value='MTTR Performance:').font = self.styles['bold_font']
            ws.cell(row=start_row+2, column=2, value=f"{mttr_score:.1f}/100").font = self.styles['normal_font']
            
            # Safely extract MTD data
            mtd_data = data.get('mtd', {})
            mtd_hours = mtd_data.get('mtd_hours', 0) if isinstance(mtd_data, dict) else 0
            mtd_score = self._calculate_performance_score(mtd_hours, Config.PERFORMANCE_THRESHOLDS['MTD'])
            
            ws.cell(row=start_row+3, column=1, value='MTD Performance:').font = self.styles['bold_font']
            ws.cell(row=start_row+3, column=2, value=f"{mtd_score:.1f}/100").font = self.styles['normal_font']
            
        except Exception as e:
            print(f"ERROR: Failed to create performance scores: {e}")
            ws.cell(row=start_row, column=1, value='Performance Scores').font = Font(name='Calibri', size=14, bold=True)
            ws.cell(row=start_row+2, column=1, value='MTTR Performance:').font = self.styles['bold_font']
            ws.cell(row=start_row+2, column=2, value='N/A').font = self.styles['normal_font']
            ws.cell(row=start_row+3, column=1, value='MTD Performance:').font = self.styles['bold_font']
            ws.cell(row=start_row+3, column=2, value='N/A').font = self.styles['normal_font']
    
    def _create_recommendations(self, ws, data: Dict[str, Any], start_row: int):
        """Create recommendations section"""
        ws.cell(row=start_row, column=1, value='Recommendations').font = Font(name='Calibri', size=14, bold=True)
        
        recommendations = [
            'Review and optimize ticket resolution workflows',
            'Implement automated triage for common alerts',
            'Set up SLA monitoring and alerting',
            'Consider 24/7 coverage for critical incidents',
            'Regular review of false positive rates'
        ]
        
        for i, rec in enumerate(recommendations, 1):
            ws.cell(row=start_row+i+1, column=1, value=f"{i}. {rec}").font = self.styles['normal_font']
    
    def _create_metrics_sheet(self, data: Dict[str, Any]):
        """Create detailed metrics sheet"""
        ws = self.wb.create_sheet(Config.EXCEL_CONFIG['SHEETS']['METRICS'])
        
        # Title
        ws['A1'] = 'Detailed Metrics Analysis'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Time metrics
        row = 3
        self._create_time_metrics_table(ws, data, row)
        
        # Resolution breakdown
        row = 12
        self._create_resolution_table(ws, data, row)
    
    def _create_trends_sheet(self, data: Dict[str, Any]):
        """Create trends analysis sheet"""
        ws = self.wb.create_sheet(Config.EXCEL_CONFIG['SHEETS']['TRENDS'])
        
        # Title
        ws['A1'] = 'Trends Analysis'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Add trend data if available
        if 'weekly_trends' in data:
            self._create_trends_chart(ws, data['weekly_trends'], 3)
    
    def _create_breakdown_sheet(self, data: Dict[str, Any]):
        """Create resolution breakdown sheet"""
        ws = self.wb.create_sheet(Config.EXCEL_CONFIG['SHEETS']['BREAKDOWN'])
        
        # Title
        ws['A1'] = 'Resolution Breakdown'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Resolution data
        resolution_data = data.get('resolution_breakdown', {})
        if resolution_data:
            self._create_resolution_chart(ws, resolution_data, 3)
    
    def _create_performance_sheet(self, data: Dict[str, Any]):
        """Create performance analysis sheet"""
        ws = self.wb.create_sheet(Config.EXCEL_CONFIG['SHEETS']['PERFORMANCE'])
        
        # Title
        ws['A1'] = 'Performance Analysis'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Performance metrics
        row = 3
        self._create_performance_table(ws, data, row)
    
    def _create_sla_sheet(self, data: Dict[str, Any]):
        """Create SLA compliance sheet"""
        ws = self.wb.create_sheet(Config.EXCEL_CONFIG['SHEETS']['SLA'])
        
        # Title
        ws['A1'] = 'SLA Compliance Analysis'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # SLA data
        row = 3
        self._create_sla_table(ws, data, row)
    
    def _create_raw_data_sheet(self, data: Dict[str, Any]):
        """Create raw data sheet"""
        ws = self.wb.create_sheet(Config.EXCEL_CONFIG['SHEETS']['RAW_DATA'])
        
        # Title
        ws['A1'] = 'Raw Data'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Add raw data if available
        if 'raw_data' in data:
            self._create_raw_data_table(ws, data['raw_data'], 3)
    
    def _create_time_metrics_table(self, ws, data: Dict[str, Any], start_row: int):
        """Create time metrics table"""
        headers = ['Metric', 'Calendar Hours', 'Working Hours', 'Calendar Days', 'Working Days']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.styles['subheader_font']
            cell.fill = self.styles['subheader_fill']
            cell.border = self.styles['border']
        
        # MTTR data
        mttr_data = data.get('mttr', {})
        row = start_row + 1
        ws.cell(row=row, column=1, value='MTTR').font = self.styles['bold_font']
        ws.cell(row=row, column=2, value=mttr_data.get('mttr_hours', 0)).font = self.styles['normal_font']
        ws.cell(row=row, column=3, value=mttr_data.get('mttr_working_hours', 0)).font = self.styles['normal_font']
        ws.cell(row=row, column=4, value=mttr_data.get('mttr_days', 0)).font = self.styles['normal_font']
        ws.cell(row=row, column=5, value=mttr_data.get('mttr_working_days', 0)).font = self.styles['normal_font']
        
        # MTD data
        mtd_data = data.get('mtd', {})
        row = start_row + 2
        ws.cell(row=row, column=1, value='MTD').font = self.styles['bold_font']
        ws.cell(row=row, column=2, value=mtd_data.get('mtd_hours', 0)).font = self.styles['normal_font']
        ws.cell(row=row, column=3, value=mtd_data.get('mtd_working_hours', 0)).font = self.styles['normal_font']
        ws.cell(row=row, column=4, value=mtd_data.get('mtd_days', 0)).font = self.styles['normal_font']
        ws.cell(row=row, column=5, value=mtd_data.get('mtd_working_days', 0)).font = self.styles['normal_font']
        
        # Apply borders
        for row in range(start_row, start_row + 3):
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.styles['border']
    
    def _create_resolution_table(self, ws, data: Dict[str, Any], start_row: int):
        """Create resolution breakdown table"""
        ws.cell(row=start_row, column=1, value='Resolution Breakdown').font = Font(name='Calibri', size=14, bold=True)
        
        headers = ['Resolution Type', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col, value=header)
            cell.font = self.styles['subheader_font']
            cell.fill = self.styles['subheader_fill']
            cell.border = self.styles['border']
        
        resolution_data = data.get('resolution_breakdown', {})
        total_tickets = data.get('total_tickets', 1)
        
        row = start_row + 2
        for resolution_type, count in resolution_data.items():
            percentage = (count / total_tickets) * 100 if total_tickets > 0 else 0
            
            ws.cell(row=row, column=1, value=resolution_type).font = self.styles['normal_font']
            ws.cell(row=row, column=2, value=count).font = self.styles['normal_font']
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").font = self.styles['normal_font']
            
            # Apply borders
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.styles['border']
            
            row += 1
    
    def _create_performance_table(self, ws, data: Dict[str, Any], start_row: int):
        """Create performance analysis table"""
        ws.cell(row=start_row, column=1, value='Performance Analysis').font = Font(name='Calibri', size=14, bold=True)
        
        headers = ['Metric', 'Current', 'Target', 'Performance Score', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col, value=header)
            cell.font = self.styles['subheader_font']
            cell.fill = self.styles['subheader_fill']
            cell.border = self.styles['border']
        
        # MTTR Performance
        mttr_hours = data.get('mttr', {}).get('mttr_hours', 0)
        mttr_score = self._calculate_performance_score(mttr_hours, Config.PERFORMANCE_THRESHOLDS['MTTR'])
        mttr_status = self._get_performance_status(mttr_score)
        
        row = start_row + 2
        ws.cell(row=row, column=1, value='MTTR').font = self.styles['bold_font']
        ws.cell(row=row, column=2, value=f"{mttr_hours:.2f} hours").font = self.styles['normal_font']
        ws.cell(row=row, column=3, value="< 4 hours").font = self.styles['normal_font']
        ws.cell(row=row, column=4, value=f"{mttr_score:.1f}/100").font = self.styles['normal_font']
        ws.cell(row=row, column=5, value=mttr_status).font = self.styles['normal_font']
        
        # MTD Performance
        mtd_hours = data.get('mtd', {}).get('mtd_hours', 0)
        mtd_score = self._calculate_performance_score(mtd_hours, Config.PERFORMANCE_THRESHOLDS['MTD'])
        mtd_status = self._get_performance_status(mtd_score)
        
        row = start_row + 3
        ws.cell(row=row, column=1, value='MTD').font = self.styles['bold_font']
        ws.cell(row=row, column=2, value=f"{mtd_hours:.2f} hours").font = self.styles['normal_font']
        ws.cell(row=row, column=3, value="< 1 hour").font = self.styles['normal_font']
        ws.cell(row=row, column=4, value=f"{mtd_score:.1f}/100").font = self.styles['normal_font']
        ws.cell(row=row, column=5, value=mtd_status).font = self.styles['normal_font']
        
        # Apply borders
        for row in range(start_row+1, start_row+4):
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.styles['border']
    
    def _create_sla_table(self, ws, data: Dict[str, Any], start_row: int):
        """Create SLA compliance table"""
        ws.cell(row=start_row, column=1, value='SLA Compliance Analysis').font = Font(name='Calibri', size=14, bold=True)
        
        headers = ['Severity', 'SLA Threshold', 'Average Time', 'Compliance Rate', 'Breaches']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col, value=header)
            cell.font = self.styles['subheader_font']
            cell.fill = self.styles['subheader_fill']
            cell.border = self.styles['border']
        
        sla_thresholds = Config.SLA_THRESHOLDS
        total_breaches = data.get('sla_breaches', 0)
        total_tickets = data.get('total_tickets', 1)
        
        row = start_row + 2
        for severity, threshold in sla_thresholds.items():
            compliance_rate = ((total_tickets - total_breaches) / total_tickets) * 100 if total_tickets > 0 else 0
            
            ws.cell(row=row, column=1, value=severity).font = self.styles['bold_font']
            ws.cell(row=row, column=2, value=f"{threshold} hours").font = self.styles['normal_font']
            ws.cell(row=row, column=3, value="N/A").font = self.styles['normal_font']  # Would need detailed data
            ws.cell(row=row, column=4, value=f"{compliance_rate:.1f}%").font = self.styles['normal_font']
            ws.cell(row=row, column=5, value=total_breaches).font = self.styles['normal_font']
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.styles['border']
            
            row += 1
    
    def _create_resolution_chart(self, ws, resolution_data: Dict[str, int], start_row: int):
        """Create resolution breakdown pie chart"""
        # Add data for chart
        row = start_row
        ws.cell(row=row, column=1, value='Resolution Type').font = self.styles['bold_font']
        ws.cell(row=row, column=2, value='Count').font = self.styles['bold_font']
        
        row += 1
        for resolution_type, count in resolution_data.items():
            ws.cell(row=row, column=1, value=resolution_type).font = self.styles['normal_font']
            ws.cell(row=row, column=2, value=count).font = self.styles['normal_font']
            row += 1
        
        # Create pie chart
        chart = PieChart()
        chart.title = "Resolution Breakdown"
        chart.height = 15
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row+len(resolution_data))
        titles = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+len(resolution_data))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(titles)
        
        ws.add_chart(chart, f"D{start_row}")
    
    def _create_trends_chart(self, ws, trends_data: Dict[str, List], start_row: int):
        """Create trends line chart"""
        # Add data for chart
        row = start_row
        ws.cell(row=row, column=1, value='Week').font = self.styles['bold_font']
        ws.cell(row=row, column=2, value='MTTR').font = self.styles['bold_font']
        ws.cell(row=row, column=3, value='MTD').font = self.styles['bold_font']
        
        # Add sample data (would need real trend data)
        for i in range(1, 5):
            ws.cell(row=row+i, column=1, value=f"Week {i}").font = self.styles['normal_font']
            ws.cell(row=row+i, column=2, value=4.0 + (i * 0.5)).font = self.styles['normal_font']
            ws.cell(row=row+i, column=3, value=1.0 + (i * 0.2)).font = self.styles['normal_font']
        
        # Create line chart
        chart = LineChart()
        chart.title = "Weekly Trends"
        chart.height = 15
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=start_row, max_col=3, max_row=start_row+4)
        cats = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+4)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f"E{start_row}")
    
    def _create_raw_data_table(self, ws, raw_data: List[Dict], start_row: int):
        """Create raw data table"""
        if not raw_data:
            ws.cell(row=start_row, column=1, value='No raw data available').font = self.styles['normal_font']
            return
        
        # Headers
        headers = list(raw_data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.styles['subheader_font']
            cell.fill = self.styles['subheader_fill']
            cell.border = self.styles['border']
        
        # Data
        for i, row_data in enumerate(raw_data, 1):
            for col, value in enumerate(row_data.values(), 1):
                safe_value = self._safe_excel_value(value)
                cell = ws.cell(row=start_row+i, column=col, value=safe_value)
                cell.font = self.styles['normal_font']
                cell.border = self.styles['border']
    
    def _get_status_color(self, value: float, target: float, reverse: bool = False) -> str:
        """Get status color based on performance"""
        if reverse:
            is_good = value <= target
        else:
            is_good = value <= target
        
        if is_good:
            return "🟢 Good"
        else:
            return "🔴 Needs Improvement"
    
    def _calculate_performance_score(self, value: float, thresholds: Dict[str, float]) -> float:
        """Calculate performance score (0-100)"""
        if value <= thresholds['excellent']:
            return 100
        elif value <= thresholds['good']:
            return 80
        elif value <= thresholds['acceptable']:
            return 60
        else:
            return 40
    
    def _get_performance_status(self, score: float) -> str:
        """Get performance status based on score"""
        if score >= 90:
            return "Excellent"
        elif score >= 70:
            return "Good"
        elif score >= 50:
            return "Acceptable"
        else:
            return "Poor" 